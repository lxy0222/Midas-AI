# -*- coding: utf-8 -*-
import os
import json
import tempfile
from typing import Dict, Optional, Tuple, List
from pathlib import Path
import asyncio
import logging

# 导入marker组件
try:
    from marker.converters.pdf import PdfConverter
    from marker.models import create_model_dict
    from marker.config.parser import ConfigParser
    from marker.output import text_from_rendered
    MARKER_AVAILABLE = True
except ImportError:
    MARKER_AVAILABLE = False
    print("⚠️  Marker未安装，使用基础文档处理模式")

logger = logging.getLogger(__name__)

class MarkerDocumentService:
    """基于Marker的通用文档分析服务 - 支持多种文档类型"""

    def __init__(self, config: Optional[Dict] = None):
        """
        初始化Marker文档服务

        Args:
            config: 配置字典，包含LLM服务配置等
        """
        # 默认配置
        self.default_config = {
            "output_format": "markdown",
            "output_dir": "output",
            "use_llm": False,  # 默认不使用LLM，避免API调用
            "disable_image_extraction": False,
            "extract_images": True,
            "max_pages": None,  # 不限制页数
        }

        # 合并用户配置
        self.config = {**self.default_config, **(config or {})}

        # 初始化配置解析器
        if MARKER_AVAILABLE:
            self.config_parser = ConfigParser(self.config)
        else:
            self.config_parser = None

        # 初始化转换器
        self._init_converters()

        # 支持的文件类型映射
        self.supported_formats = {
            '.pdf': 'pdf',
            '.txt': 'text',
            '.md': 'text',
            '.json': 'text',
            '.csv': 'text',
            '.log': 'text',
            '.py': 'text',
            '.js': 'text',
            '.html': 'text',
            '.css': 'text',
            '.java': 'text',
            '.cpp': 'text',
            '.c': 'text',
            '.go': 'text',
            '.rs': 'text',
            '.docx': 'office',  # 通过转换为PDF处理
            '.doc': 'office',
            '.xlsx': 'office',
            '.xls': 'office',
            '.pptx': 'office',
            '.ppt': 'office'
        }
    
    def _init_converters(self):
        """初始化各种文档转换器"""
        if not MARKER_AVAILABLE:
            logger.warning("Marker未安装，转换器将不可用")
            self.pdf_converter = None
            return

        try:
            # 创建模型字典
            artifact_dict = create_model_dict()

            # 获取处理器列表
            processor_list = self.config_parser.get_processors()

            # 获取渲染器
            renderer = self.config_parser.get_renderer()

            # 创建PDF转换器
            self.pdf_converter = PdfConverter(
                config=self.config,
                artifact_dict=artifact_dict,
                processor_list=processor_list,
                renderer=renderer,
            )

            logger.info("Marker转换器初始化成功")

        except Exception as e:
            logger.error(f"初始化Marker转换器失败: {str(e)}")
            self.pdf_converter = None
    
    def enable_llm_service(self, llm_config: Dict):
        """
        启用LLM服务
        
        Args:
            llm_config: LLM配置，包含服务类型、API密钥等
        """
        try:
            # 更新配置
            self.config.update({
                "use_llm": True,
                **llm_config
            })
            
            # 重新初始化转换器
            self._init_converters()
            
            logger.info("LLM服务已启用")
            
        except Exception as e:
            logger.error(f"启用LLM服务失败: {str(e)}")
    
    async def extract_document_content(self, file_path: str, filename: str = None) -> Dict:
        """
        使用Marker提取文档内容 - 支持多种文档类型

        Args:
            file_path: 文件路径
            filename: 文件名（用于确定文件类型）

        Returns:
            包含提取结果的字典
        """
        if filename is None:
            filename = os.path.basename(file_path)

        # 获取文件类型
        file_ext = Path(filename).suffix.lower()
        file_type = self.supported_formats.get(file_ext, 'unknown')

        if file_type == 'unknown':
            return {
                'success': False,
                'error': f'不支持的文件类型: {file_ext}',
                'content': '',
                'metadata': {},
                'images': {}
            }

        try:
            # 检查文件是否存在
            abs_path = os.path.abspath(file_path)
            if not os.path.exists(abs_path):
                return {
                    'success': False,
                    'error': f'文件不存在: {abs_path}',
                    'content': '',
                    'metadata': {},
                    'images': {}
                }

            # 根据文件类型选择处理方法
            if file_type == 'pdf':
                return await self._process_pdf(abs_path, filename)
            elif file_type == 'text':
                return await self._process_text_file(abs_path, filename)
            elif file_type == 'office':
                return await self._process_office_file(abs_path, filename)
            else:
                return {
                    'success': False,
                    'error': f'暂不支持的文件类型: {file_type}',
                    'content': '',
                    'metadata': {},
                    'images': {}
                }

        except Exception as e:
            logger.error(f"Marker文档提取失败: {str(e)}")
            return {
                'success': False,
                'error': f'文档处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }
    
    async def _process_pdf(self, file_path: str, filename: str) -> Dict:
        """处理PDF文件"""
        if not MARKER_AVAILABLE or not self.pdf_converter:
            # 使用基础PDF处理
            return await self._fallback_pdf_processing(file_path, filename)

        try:
            # 在线程池中运行转换（因为marker是同步的）
            loop = asyncio.get_event_loop()
            rendered = await loop.run_in_executor(
                None,
                self._convert_pdf_sync,
                file_path
            )

            if rendered is None:
                return {
                    'success': False,
                    'error': 'PDF转换失败',
                    'content': '',
                    'metadata': {},
                    'images': {}
                }

            # 提取文本和图像
            if MARKER_AVAILABLE:
                text, metadata, images = text_from_rendered(rendered)
            else:
                # 这个分支不应该被执行，因为前面已经检查了MARKER_AVAILABLE
                text = "Marker处理失败"
                metadata = {}
                images = {}

            # 处理图像引用
            processed_text = self._process_image_references(text, images)

            # 构建元数据
            doc_metadata = {
                'type': 'pdf_marker',
                'pages': metadata.get('pages', 0) if metadata else 0,
                'size': len(processed_text),
                'images_count': len(images),
                'extraction_method': 'marker',
                'llm_enabled': self.config.get('use_llm', False),
                'filename': filename
            }

            return {
                'success': True,
                'content': processed_text,
                'metadata': doc_metadata,
                'images': images
            }

        except Exception as e:
            logger.error(f"PDF处理失败: {str(e)}")
            return {
                'success': False,
                'error': f'PDF处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    def _convert_pdf_sync(self, file_path: str):
        """同步转换PDF文件"""
        try:
            return self.pdf_converter(file_path)
        except Exception as e:
            logger.error(f"PDF转换过程中出错: {str(e)}")
            return None

    async def _process_text_file(self, file_path: str, filename: str) -> Dict:
        """处理文本文件 - 使用Marker的智能格式化"""
        try:
            # 读取文本内容
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # 对于代码文件，保持原格式并添加语法高亮标记
            file_ext = Path(filename).suffix.lower()
            if file_ext in ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.go', '.rs']:
                language = self._get_language_from_extension(file_ext)
                processed_content = f"```{language}\n{content}\n```"
            else:
                # 对于普通文本文件，使用Marker的智能格式化
                processed_content = self._format_text_content(content, file_ext)

            # 构建元数据
            doc_metadata = {
                'type': 'text_marker',
                'size': len(processed_content),
                'lines': len(content.split('\n')),
                'extraction_method': 'marker_text',
                'language': self._get_language_from_extension(file_ext),
                'filename': filename
            }

            return {
                'success': True,
                'content': processed_content,
                'metadata': doc_metadata,
                'images': {}
            }

        except UnicodeDecodeError:
            # 尝试其他编码
            try:
                with open(file_path, 'r', encoding='gbk') as f:
                    content = f.read()

                processed_content = self._format_text_content(content, Path(filename).suffix.lower())

                return {
                    'success': True,
                    'content': processed_content,
                    'metadata': {
                        'type': 'text_marker',
                        'encoding': 'gbk',
                        'size': len(processed_content),
                        'lines': len(content.split('\n')),
                        'extraction_method': 'marker_text',
                        'filename': filename
                    },
                    'images': {}
                }
            except Exception as e:
                return {
                    'success': False,
                    'error': f'无法读取文件编码: {str(e)}',
                    'content': '',
                    'metadata': {},
                    'images': {}
                }
        except Exception as e:
            return {
                'success': False,
                'error': f'文本文件处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    async def _process_office_file(self, file_path: str, filename: str) -> Dict:
        """处理Office文件 - 通过转换为PDF后使用Marker处理"""
        try:
            # 这里可以集成LibreOffice或其他工具将Office文件转换为PDF
            # 然后使用Marker的PDF处理功能
            # 目前返回一个占位符实现

            file_ext = Path(filename).suffix.lower()

            if file_ext in ['.docx', '.doc']:
                return await self._process_word_file(file_path, filename)
            elif file_ext in ['.xlsx', '.xls']:
                return await self._process_excel_file(file_path, filename)
            elif file_ext in ['.pptx', '.ppt']:
                return await self._process_powerpoint_file(file_path, filename)
            else:
                return {
                    'success': False,
                    'error': f'不支持的Office文件类型: {file_ext}',
                    'content': '',
                    'metadata': {},
                    'images': {}
                }

        except Exception as e:
            return {
                'success': False,
                'error': f'Office文件处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    def _process_image_references(self, text: str, images: Dict) -> str:
        """
        处理文本中的图像引用
        
        Args:
            text: 原始文本
            images: 图像字典
            
        Returns:
            处理后的文本
        """
        processed_text = text
        
        # 替换图像引用为描述性文本
        for image_key in images.keys():
            image_ref = f"![]({image_key})"
            if image_ref in processed_text:
                processed_text = processed_text.replace(
                    image_ref, 
                    f"[图像: {image_key}]"
                )
        
        return processed_text

    def _get_language_from_extension(self, ext: str) -> str:
        """根据文件扩展名获取编程语言"""
        language_map = {
            '.py': 'python',
            '.js': 'javascript',
            '.html': 'html',
            '.css': 'css',
            '.java': 'java',
            '.cpp': 'cpp',
            '.c': 'c',
            '.go': 'go',
            '.rs': 'rust',
            '.json': 'json',
            '.md': 'markdown',
            '.txt': 'text'
        }
        return language_map.get(ext, 'text')

    def _format_text_content(self, content: str, file_ext: str) -> str:
        """格式化文本内容"""
        if file_ext == '.json':
            try:
                # 格式化JSON
                import json
                parsed = json.loads(content)
                return f"```json\n{json.dumps(parsed, indent=2, ensure_ascii=False)}\n```"
            except:
                return f"```json\n{content}\n```"
        elif file_ext == '.md':
            # Markdown文件直接返回
            return content
        elif file_ext == '.csv':
            # CSV文件添加表格格式
            lines = content.split('\n')
            if len(lines) > 1:
                # 简单的CSV到Markdown表格转换
                formatted_lines = []
                for i, line in enumerate(lines[:10]):  # 只显示前10行
                    if line.strip():
                        formatted_lines.append('| ' + ' | '.join(line.split(',')) + ' |')
                        if i == 0:  # 添加表头分隔符
                            formatted_lines.append('| ' + ' | '.join(['---'] * len(line.split(','))) + ' |')
                return '\n'.join(formatted_lines)
            return content
        else:
            return content

    async def _process_word_file(self, file_path: str, filename: str) -> Dict:
        """处理Word文档"""
        try:
            from docx import Document
            doc = Document(file_path)

            paragraphs = []
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text)

            content = '\n\n'.join(paragraphs)

            # 使用Marker风格的格式化
            formatted_content = f"# {filename}\n\n{content}"

            return {
                'success': True,
                'content': formatted_content,
                'metadata': {
                    'type': 'docx_marker',
                    'paragraphs': len(paragraphs),
                    'size': len(formatted_content),
                    'extraction_method': 'marker_docx',
                    'filename': filename
                },
                'images': {}
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'Word文档处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    async def _process_excel_file(self, file_path: str, filename: str) -> Dict:
        """处理Excel文件"""
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True)

            sheets_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []

                # 转换为Markdown表格格式
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=50)):  # 限制行数
                    if any(cell is not None for cell in row):
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        sheet_data.append('| ' + ' | '.join(row_data) + ' |')
                        if row_idx == 0:  # 添加表头分隔符
                            sheet_data.append('| ' + ' | '.join(['---'] * len(row_data)) + ' |')

                if sheet_data:
                    sheets_content.append(f"## {sheet_name}\n\n" + '\n'.join(sheet_data))

            content = f"# {filename}\n\n" + '\n\n'.join(sheets_content)

            return {
                'success': True,
                'content': content,
                'metadata': {
                    'type': 'excel_marker',
                    'sheets': len(workbook.sheetnames),
                    'size': len(content),
                    'extraction_method': 'marker_excel',
                    'filename': filename
                },
                'images': {}
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel文件处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    async def _process_powerpoint_file(self, file_path: str, filename: str) -> Dict:
        """处理PowerPoint文件"""
        try:
            # 这里可以集成python-pptx或其他库
            # 目前返回占位符
            return {
                'success': False,
                'error': 'PowerPoint文件处理功能正在开发中',
                'content': '',
                'metadata': {},
                'images': {}
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'PowerPoint文件处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }

    def get_supported_formats(self) -> List[str]:
        """获取支持的文件格式"""
        return list(self.supported_formats.keys())

    def is_supported_file(self, filename: str) -> bool:
        """检查文件是否支持"""
        ext = Path(filename).suffix.lower()
        return ext in self.supported_formats

    async def _fallback_pdf_processing(self, file_path: str, filename: str) -> Dict:
        """当Marker不可用时的PDF基础处理"""
        try:
            import PyPDF2

            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = []

                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        text = page.extract_text()
                        if text.strip():
                            text_content.append(f"=== 第 {page_num + 1} 页 ===\n{text}")
                    except Exception as page_error:
                        text_content.append(f"=== 第 {page_num + 1} 页 ===\n[页面解析失败: {str(page_error)}]")

                content = '\n\n'.join(text_content)

                return {
                    'success': True,
                    'content': content,
                    'metadata': {
                        'type': 'pdf_basic',
                        'pages': len(pdf_reader.pages),
                        'size': len(content),
                        'extraction_method': 'pypdf2_fallback',
                        'filename': filename
                    },
                    'images': {}
                }
        except Exception as e:
            return {
                'success': False,
                'error': f'PDF基础处理失败: {str(e)}',
                'content': '',
                'metadata': {},
                'images': {}
            }


# 创建全局实例
marker_service = MarkerDocumentService()
