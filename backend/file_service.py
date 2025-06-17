import os
import io
import json
import time
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import aiofiles
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import base64

class FileService:
    """文件处理服务"""
    
    def __init__(self):
        self.upload_dir = Path("uploads")
        self.upload_dir.mkdir(exist_ok=True)
        
        # 支持的文件类型
        self.supported_types = {
            'text': ['.txt', '.md', '.json', '.csv', '.log'],
            'document': ['.pdf', '.docx', '.doc'],
            'spreadsheet': ['.xlsx', '.xls'],
            'image': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'],
            'code': ['.py', '.js', '.html', '.css', '.java', '.cpp', '.c', '.go', '.rs']
        }
        
        # 文件大小限制 (MB)
        self.max_file_size = 10
    
    def get_file_type(self, filename: str) -> str:
        """获取文件类型"""
        ext = Path(filename).suffix.lower()
        for file_type, extensions in self.supported_types.items():
            if ext in extensions:
                return file_type
        return 'unknown'
    
    def is_supported_file(self, filename: str) -> bool:
        """检查是否支持该文件类型"""
        return self.get_file_type(filename) != 'unknown'
    
    async def save_uploaded_file(self, file_content: bytes, filename: str) -> str:
        """保存上传的文件"""
        # 生成唯一文件名
        timestamp = str(int(time.time() * 1000))
        safe_filename = f"{timestamp}_{filename}"
        file_path = self.upload_dir / safe_filename

        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(file_content)

        return str(file_path)
    
    async def extract_text_from_file(self, file_path: str, filename: str) -> Dict:
        """从文件中提取文本内容"""
        file_type = self.get_file_type(filename)
        
        try:
            if file_type == 'text' or file_type == 'code':
                return await self._extract_text_file(file_path)
            elif file_type == 'document':
                if filename.lower().endswith('.pdf'):
                    return await self._extract_pdf(file_path)
                elif filename.lower().endswith(('.docx', '.doc')):
                    return await self._extract_docx(file_path)
            elif file_type == 'spreadsheet':
                return await self._extract_excel(file_path)
            elif file_type == 'image':
                return await self._extract_image_info(file_path)
            else:
                return {
                    'success': False,
                    'error': f'不支持的文件类型: {file_type}',
                    'content': '',
                    'metadata': {}
                }
        except Exception as e:
            return {
                'success': False,
                'error': f'文件处理失败: {str(e)}',
                'content': '',
                'metadata': {}
            }
    
    async def _extract_text_file(self, file_path: str) -> Dict:
        """提取文本文件内容"""
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                content = await f.read()
            
            return {
                'success': True,
                'content': content,
                'metadata': {
                    'type': 'text',
                    'size': len(content),
                    'lines': len(content.split('\n'))
                }
            }
        except UnicodeDecodeError:
            # 尝试其他编码
            try:
                async with aiofiles.open(file_path, 'r', encoding='gbk') as f:
                    content = await f.read()
                return {
                    'success': True,
                    'content': content,
                    'metadata': {
                        'type': 'text',
                        'encoding': 'gbk',
                        'size': len(content),
                        'lines': len(content.split('\n'))
                    }
                }
            except:
                return {
                    'success': False,
                    'error': '无法读取文件编码',
                    'content': '',
                    'metadata': {}
                }
    
    async def _extract_pdf(self, file_path: str) -> Dict:
        """提取PDF文件内容"""
        try:
            # 使用绝对路径确保文件存在
            abs_path = os.path.abspath(file_path)
            if not os.path.exists(abs_path):
                return {
                    'success': False,
                    'error': 'PDF文件不存在',
                    'content': '',
                    'metadata': {}
                }

            with open(abs_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = []

                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        text = page.extract_text()
                        if text.strip():
                            text_content.append(f"=== 第 {page_num + 1} 页 ===\n{text}")
                    except Exception as page_error:
                        text_content.append(f"=== 第 {page_num + 1} 页 ===\n[页面解析失败: {str(page_error)}]")

                content = '\n\n'.join(text_content)

                return {
                    'success': True,
                    'content': content,
                    'metadata': {
                        'type': 'pdf',
                        'pages': len(pdf_reader.pages),
                        'size': len(content)
                    }
                }
        except Exception as e:
            return {
                'success': False,
                'error': f'PDF处理失败: {str(e)}',
                'content': '',
                'metadata': {}
            }
    
    async def _extract_docx(self, file_path: str) -> Dict:
        """提取Word文档内容"""
        try:
            doc = Document(file_path)
            paragraphs = []
            
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text)
            
            content = '\n\n'.join(paragraphs)
            
            return {
                'success': True,
                'content': content,
                'metadata': {
                    'type': 'docx',
                    'paragraphs': len(paragraphs),
                    'size': len(content)
                }
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'Word文档处理失败: {str(e)}',
                'content': '',
                'metadata': {}
            }
    
    async def _extract_excel(self, file_path: str) -> Dict:
        """提取Excel文件内容"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheets_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        sheet_data.append('\t'.join(row_data))
                
                if sheet_data:
                    sheets_content.append(f"=== 工作表: {sheet_name} ===\n" + '\n'.join(sheet_data))
            
            content = '\n\n'.join(sheets_content)
            
            return {
                'success': True,
                'content': content,
                'metadata': {
                    'type': 'excel',
                    'sheets': len(workbook.sheetnames),
                    'size': len(content)
                }
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel文件处理失败: {str(e)}',
                'content': '',
                'metadata': {}
            }
    
    async def _extract_image_info(self, file_path: str) -> Dict:
        """提取图片信息"""
        try:
            abs_path = os.path.abspath(file_path)
            if not os.path.exists(abs_path):
                return {
                    'success': False,
                    'error': '图片文件不存在',
                    'content': '',
                    'metadata': {}
                }

            # 先读取图片信息
            with Image.open(abs_path) as img:
                width, height = img.size
                format_name = img.format
                mode = img.mode

            # 单独读取文件数据用于base64编码
            with open(abs_path, 'rb') as f:
                img_data = f.read()
                img_base64 = base64.b64encode(img_data).decode('utf-8')

            content = f"这是一张图片文件，格式：{format_name}，尺寸：{width}x{height}像素，颜色模式：{mode}"

            return {
                'success': True,
                'content': content,
                'metadata': {
                    'type': 'image',
                    'format': format_name,
                    'width': width,
                    'height': height,
                    'mode': mode,
                    'base64': f"data:image/{format_name.lower()};base64,{img_base64}"
                }
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'图片处理失败: {str(e)}',
                'content': '',
                'metadata': {}
            }
    
    async def cleanup_old_files(self, max_age_hours: int = 24):
        """清理旧文件"""
        current_time = time.time()

        for file_path in self.upload_dir.glob('*'):
            if file_path.is_file():
                file_age = current_time - file_path.stat().st_mtime
                if file_age > max_age_hours * 3600:  # 转换为秒
                    try:
                        file_path.unlink()
                    except:
                        pass

# 创建全局实例
file_service = FileService()
